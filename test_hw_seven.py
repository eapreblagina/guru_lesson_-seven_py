
from PyPDF2 import PdfReader
from xlrd import open_workbook
from openpyxl import load_workbook
from zipfile import ZipFile

#
# PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
# RESOURCES_FOLDER = os.path.join(PROJECT_ROOT_PATH, 'resources')
# TMP_FOLDER = os.path.join(PROJECT_ROOT_PATH, 'tmp')
#
# archive_name = 'archive.zip'
#
# txt_file = os.path.join(RESOURCES_FOLDER, 'file.txt')
# xls_file = os.path.join(RESOURCES_FOLDER, 'book.xls')
# xlsx_file = os.path.join(RESOURCES_FOLDER, 'тест-кейс пример.xlsx')
# pdf_file = os.path.join(RESOURCES_FOLDER, 'Екатерина Преблагина.pdf')


# def test_zip_files():
#     if not os.path.exists(TMP_FOLDER):
#         os.mkdir(TMP_FOLDER)
#
#     path_to_archive = f'{TMP_FOLDER}/{archive_name}'
#
#     with zipfile.ZipFile(path_to_archive, 'w') as file:
#         file.write(txt_file, basename(txt_file))
#         file.write(xls_file, basename(xls_file))
#         file.write(xlsx_file, basename(xlsx_file))
#         file.write(pdf_file, basename(pdf_file))

# def test_unzip_files(test_zip_files):
#     with ZipFile("tmp/archive.zip") as zip_file:
#         print(zip_file.filelist)
#         test_xlsx_file()
#         test_pdf_file()
#         test_xls_file()
#         test_txt_file()

def test_xls_file():
    with ZipFile("tmp/archive.zip") as zip_file:
            book = open_workbook(file_contents=zip_file.read("book.xls"))
            print(f"Количество листов {book.nsheets}")
            print(f"Имена листов{book.sheet_names()}")
            sheet = book.sheet_by_index(0)
            print(sheet.nrows)
            print(sheet.ncols)
            print(sheet.cell_value(2, 3))
            for rx in range(sheet.nrows):
                print(sheet.row(rx))


def test_pdf_file():
    with ZipFile("tmp/archive.zip") as zip_file:
        with zip_file.open("Екатерина Преблагина.pdf") as file:
            reader = PdfReader(file)
            number_of_pages = len(reader.pages)
            print(number_of_pages)
            page = reader.pages[1]
            text = page.extract_text()
            print(text)


def test_xlsx_file():
    with ZipFile("tmp/archive.zip") as zip_file:
        with zip_file.open("тест-кейс пример.xlsx") as file:
            workbook = load_workbook(file)
            sheet = workbook.active
            print(sheet.cell(row=3, column=2).value)


def test_txt_file():
    with ZipFile("tmp/archive.zip") as zip_file:
        print(zip_file.namelist())
        text = zip_file.read('file.txt')
        print(text)
